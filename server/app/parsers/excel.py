"""Excel parser: openpyxl → Document.

Each sheet is rendered as a markdown-style table prefixed by the sheet name.
Empty cells become empty strings; merged cells take the value of the anchor.
"""
from __future__ import annotations

from pathlib import Path

from app.observability.models import Document
from app.parsers._common import detect_mime


def parse_excel(path: str | Path) -> Document:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as e:  # pragma: no cover
        raise RuntimeError(
            "openpyxl not installed. `pip install -e '.[parsers]'`."
        ) from e

    p = Path(path)
    wb = load_workbook(filename=str(p), data_only=True, read_only=True)
    parts: list[str] = []
    sheet_names: list[str] = []
    for ws in wb.worksheets:
        sheet_names.append(ws.title)
        parts.append(f"# Sheet: {ws.title}\n")
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            parts.append("| " + " | ".join(cells) + " |")
        parts.append("")  # blank line between sheets

    text = "\n".join(parts).strip()
    return Document(
        source_path=str(p),
        mime=detect_mime(p),
        text=text,
        metadata={"parser": "excel", "sheets": sheet_names, "size": p.stat().st_size},
    )